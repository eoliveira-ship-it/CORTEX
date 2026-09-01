"""Gera diag_divergencias.sql : mostra valor da TABELA vs valor do FICHEIRO,
lado a lado, para as colunas que divergiram no teste_conteudo.

Sem isto so sabemos QUE diverge; com isto sabe-se PORQUE:
  - valores parecidos mas diferentes -> dados mudaram (outra fotografia)
  - valor de outro campo             -> mapeamento errado
  - escala/arredondamento            -> problema de precisao
"""
import io
import os
import re
import sys

import openpyxl

REAL = (r'C:/Users/elder/AppData/Local/Temp/claude/'
        r'C--Users-elder-Documents/c833d650-7531-40f4-951e-02752213f565/'
        r'scratchpad/CRRCORP_P1.txt')
NL = chr(10)
Q = chr(39)

# colunas que divergiram (lidas do resultado que veio do Oracle)
DIVERG = []
if os.path.exists('testes'):
    for l in open('testes', encoding='utf-8'):
        if '_C' + chr(9) not in l:
            continue
        for c in l.split(chr(9), 1)[1].split():
            c = c.replace('(esperado', '')
            if c.startswith('P1') and c not in DIVERG:
                DIVERG.append(c)

ws = openpyxl.load_workbook('Notice PACTV4.5_v1.0.xlsx', data_only=True)['PACT Corp']
NOME, DEC = {}, {}
for r in range(4, ws.max_row + 1):
    if str(ws.cell(r, 1).value or '').strip() != 'P1':
        continue
    ref = str(ws.cell(r, 5).value or '').strip()
    NOME[ref] = str(ws.cell(r, 6).value or '').replace(NL, ' ').strip()
    m = re.search(r'(\d+)\s*d.cimale', str(ws.cell(r, 21).value or ''), re.I)
    DEC[ref] = int(m.group(1)) if m else 0

src = open('align_v44.py', encoding='utf-8').read()
src = src.split('# ------------------------------------------------------- alinhamento')[0]
ns = {}
buf = io.StringIO()
_o = sys.stdout
sys.stdout = buf
exec(src, ns)
sys.stdout = _o
tokenize, width, v44 = ns['tokenize'], ns['width'], ns['v44']


def col_de(ref):
    if '(P1)' in ref:
        return 'P1_H_' + ref.replace('(P1)', '').strip().replace('.', '_')
    return 'P1_' + ref.split()[1].replace('.', '_')


POS = {}
pos = 0
for t in tokenize(589, 1068):
    w = width(t['raw'])
    if w is None:
        nf = next((f for f in v44 if f['start'] == pos), None)
        w = nf['len'] if nf else 0
    ach = [f for f in v44 if f['start'] < pos + w and f['start'] + f['len'] > pos]
    if len(ach) == 1:
        POS.setdefault(col_de(ach[0]['ref']), (pos, w, ach[0]['ref']))
    pos += w

ID_OFF, ID_W = 180, 40
linha = None
with open(REAL, 'r', encoding='latin-1') as f:
    for l in f:
        l = l.rstrip('\r\n')
        if l[ID_OFF:ID_OFF + ID_W].strip():
            linha = l
            break
idv = linha[ID_OFF:ID_OFF + ID_W].strip()

sel = []
for c in DIVERG:
    if c not in POS:
        continue
    off, w, ref = POS[c]
    bruto = linha[off:off + w]
    sel.append("  SELECT %s%-12s%s AS coluna, %s%s%s AS no_ficheiro,"
               % (Q, c, Q, Q, bruto.strip() or ' ', Q)
               + NL + "         TO_CHAR(%s) AS na_tabela, %s%s%s AS descricao"
               % (c, Q, NOME.get(ref, '')[:38].replace(Q, ''), Q)
               + NL + "    FROM ENG_CORP_P1_BIS WHERE P1_H_1_11 = %s%s%s" % (Q, idv, Q))

sql = [
    "-- Valor da TABELA vs valor do FICHEIRO, para as colunas divergentes.",
    "-- Engajamento analisado: " + idv,
    "--",
    "-- Como ler:",
    "--   valores parecidos mas diferentes -> os dados mudaram (outra fotografia)",
    "--   valor de um campo COMPLETAMENTE outro -> mapeamento errado",
    "--   mesmo numero com virgula noutro sitio -> escala errada",
    "SET LINESIZE 200",
    "COLUMN coluna      FORMAT A12",
    "COLUMN no_ficheiro FORMAT A22",
    "COLUMN na_tabela   FORMAT A22",
    "COLUMN descricao   FORMAT A40",
    "",
    "-- 0) A fotografia e a mesma? (no ficheiro: 20250630)",
    "SELECT TO_CHAR(P1_H_0_1,'YYYYMMDD') AS arrete_na_tabela, COUNT(*) AS linhas",
    "  FROM ENG_CORP_P1_BIS GROUP BY TO_CHAR(P1_H_0_1,'YYYYMMDD') ORDER BY 1;",
    "",
    "-- 1) Valores lado a lado",
    (NL + "  UNION ALL" + NL).join(sel) + ";",
]
open('diag_divergencias.sql', 'w', encoding='utf-8').write(NL.join(sql) + NL)
print('colunas divergentes analisadas : %d' % len(sel))
print('engajamento : %s' % idv)
print('-> diag_divergencias.sql')
