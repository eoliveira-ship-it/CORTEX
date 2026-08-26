"""Gera o mapeamento posicao -> coluna para TODAS as posicoes dos 8 SELECT.

Metodo:
  1. regua V44 (align_v44.py) da o campo da notice em cada offset;
  2. o ficheiro CRRCORP_P1 real serve de oraculo: extrai-se o valor em cada
     posicao e verifica-se se bate com o FORMAT declarado na notice.

Confianca:
  ALTA   ancora '--P1' no spool E a posicao concordam
  MEDIA  sem ancora, um so campo na posicao, e os valores reais respeitam
         o formato declarado (DATE -> 8 digitos, NUM -> sinal+digitos)
  BAIXA  ancora e posicao discordam, ou a posicao cobre varios campos,
         ou os valores reais nao respeitam o formato

Saida: docs/mapa-posicoes.csv  +  resumo no ecran.
"""
import io
import os
import re
import sys

REAL = (r'C:/Users/elder/AppData/Local/Temp/claude/'
        r'C--Users-elder-Documents/c833d650-7531-40f4-951e-02752213f565/'
        r'scratchpad/CRRCORP_P1.txt')
N_AMOSTRA = 300

src = open('align_v44.py', encoding='utf-8').read()
src = src.split('# ------------------------------------------------------- alinhamento')[0]
ns = {}
buf = io.StringIO()
_o = sys.stdout
sys.stdout = buf
exec(src, ns)
sys.stdout = _o
tokenize, width, v44, notice = ns['tokenize'], ns['width'], ns['v44'], ns['notice']

FMT = {f['ref']: f['fmt'] for f in
       [dict(ref=n['ref'], fmt=n.get('fmt', '')) for n in notice]}
# a notice lida em align_v44 nao traz 'fmt' -> reler
import openpyxl
_ws = openpyxl.load_workbook('Notice PACTV4.5_v1.0.xlsx', data_only=True)['PACT Corp']
FMT = {}
NOME = {}
for r in range(4, _ws.max_row + 1):
    if str(_ws.cell(r, 1).value or '').strip() != 'P1':
        continue
    ref = str(_ws.cell(r, 5).value or '').strip()
    FMT[ref] = str(_ws.cell(r, 20).value or '').strip().upper()
    NOME[ref] = str(_ws.cell(r, 6).value or '').replace('\n', ' ').strip()


def col_de(ref):
    if '(P1)' in ref:
        return 'P1_H_' + ref.replace('(P1)', '').strip().replace('.', '_')
    return 'P1_' + ref.split()[1].replace('.', '_')


amostras = []
if os.path.exists(REAL):
    with open(REAL, 'r', encoding='latin-1') as f:
        for i, l in enumerate(f):
            amostras.append(l.rstrip('\r\n'))
            if i >= N_AMOSTRA:
                break


def confere(off, w, ref):
    """Confronta os valores reais nessa posicao com o FORMAT da notice."""
    if not amostras or w == 0:
        return None, []
    vals = {l[off:off + w] for l in amostras}
    nao_vazios = [v for v in vals if v.strip()]
    if not nao_vazios:
        return 'vazio', []
    fmt = FMT.get(ref, '')
    ok = True
    if fmt == 'DATE':
        ok = all(re.fullmatch(r'\d{8}', v.strip()) for v in nao_vazios)
    elif fmt == 'NUM':
        ok = all(re.fullmatch(r'[+-]?[\d.,]+', v.strip()) for v in nao_vazios)
    return ('ok' if ok else 'incoerente'), sorted(nao_vazios)[:2]


VAR = [(1, 589, 1068), (2, 1088, 1575), (3, 1591, 2069), (4, 2893, 3449),
       (5, 3461, 4010), (6, 4025, 4593), (7, 4605, 5049), (8, 5060, 5628)]

linhas = []
stats = {'ALTA': 0, 'MEDIA': 0, 'BAIXA': 0, 'FILLER': 0}
for num, a, b in VAR:
    toks = tokenize(a, b)
    pos = 0
    seq = 0
    for t in toks:
        w = width(t['raw'])
        if w is None:
            nf = next((f for f in v44 if f['start'] == pos), None)
            w = nf['len'] if nf else 0
        campos = [f for f in v44 if f['start'] < pos + w and f['start'] + f['len'] > pos]
        refs = [f['ref'] for f in campos]
        raw = re.sub(r'\s+', ' ', t['raw']).strip()
        e_filler = bool(re.match(r"^RPAD\s*\(\s*'\s*'\s*,", raw, re.I))
        if not e_filler:
            seq += 1
        anc = t.get('ref')
        ref0 = refs[0] if refs else ''
        estado, ex = confere(pos, w, ref0)

        if e_filler:
            conf = 'FILLER'
        elif anc:
            pref = anc + '.'
            conf = 'ALTA' if (anc in refs or any(c.startswith(pref) for c in refs)) else 'BAIXA'
        elif len(refs) == 1 and estado in ('ok', 'vazio'):
            conf = 'MEDIA'
        else:
            conf = 'BAIXA'
        stats[conf] += 1
        linhas.append([
            num, seq if not e_filler else '', pos, w,
            col_de(ref0) if ref0 else '', ref0, FMT.get(ref0, ''),
            NOME.get(ref0, '')[:45], anc or '', conf, estado or '',
            ' / '.join(ex), raw[:70],
        ])
        pos += w

hdr = ['insert', 'seq', 'offset', 'largura', 'coluna', 'ref_notice', 'formato',
       'nome_notice', 'ancora_spool', 'confianca', 'oraculo', 'exemplos', 'expressao_spool']
import csv
with open('docs/mapa-posicoes.csv', 'w', newline='', encoding='utf-8-sig') as f:
    wr = csv.writer(f, delimiter=';')
    wr.writerow(hdr)
    wr.writerows(linhas)

print('posicoes analisadas : %d' % len(linhas))
for k in ('ALTA', 'MEDIA', 'BAIXA', 'FILLER'):
    print('   %-7s %d' % (k, stats[k]))
uteis = [l for l in linhas if l[9] in ('ALTA', 'MEDIA')]
print('posicoes mapeaveis  : %d (%d colunas distintas)'
      % (len(uteis), len({l[4] for l in uteis if l[4]})))
print('-> docs/mapa-posicoes.csv')
