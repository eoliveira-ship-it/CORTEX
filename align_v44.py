"""Reconstrucao da regua V44.02 e alinhamento posicional com o spool.

IDEIA (sugerida pelo utilizador):
  - a coluna LENGTH da notice define as posicoes (soma acumulada);
  - a coluna USAGE identifica os campos nao utilizados (NA): quando estao
    em sequencia, o spool junta-os num unico RPAD(' ', soma).

REGUA V44 = notice V45 menos os campos criados em V45 (coluna VERSION DE
            CREATION = 45), com P1 21.65 revertido de 50 para 5 (SIRL-1223).

RESULTADO ATUAL
  largura total : spool 5682 vs regua V44 5675  -> delta +7 octetos
  validacao     : 92 de 120 ancoras '--P1' caem no campo certo (77%)

Os 28 restantes sao desvios de UM campo, concentrados em 4 zonas: cada um
deve corresponder a um campo cujo COMPRIMENTO mudou em V45 e que nao
sabemos reverter (so conhecemos o 21.65, dado pelo SIRL-1223).

=> Com a notice V44.02 o alinhamento fica exato e as 922 posicoes podem
   ser mapeadas automaticamente. Sem ela, isto produz uma PROPOSTA de
   mapeamento a validar, nao uma prova.

Uso:  python align_v44.py
"""
import re
import openpyxl

# ----------------------------------------------------------------- notice
ws = openpyxl.load_workbook('Notice PACTV4.5_v1.0.xlsx', data_only=True)['PACT Corp']
notice = []
for r in range(4, ws.max_row + 1):
    if str(ws.cell(r, 1).value or '').strip() != 'P1':
        continue
    e = str(ws.cell(r, 5).value or '').strip()
    try:
        ln = int(ws.cell(r, 23).value)
    except Exception:
        ln = 0
    notice.append({
        'ref': e,
        'len': ln,
        'usage': str(ws.cell(r, 28).value or '').strip(),
        'crea': str(ws.cell(r, 25).value or '').strip(),
    })

v44 = [f.copy() for f in notice if not f['crea'].startswith('45')]
for f in v44:
    if f['ref'] == 'P1 21.65':
        f['len'] = 5            # SIRL-1223 : 5 -> 50 en V45
off = 0
for f in v44:
    f['start'] = off
    off += f['len']
TOT_V44 = off
print('notice V45 : %d campos, %d octetos' % (len(notice), sum(f['len'] for f in notice)))
print('regua V44  : %d campos, %d octetos' % (len(v44), TOT_V44))

# ----------------------------------------------------------------- spool
lines = open('030_spool_Extract_CRRCORP-antigo.sql', encoding='utf-8', errors='replace').read().split('\n')


def strip_line(l, in_block):
    code, cm, i, in_str = '', '', 0, False
    while i < len(l):
        two = l[i:i + 2]
        if in_block:
            if two == '*/':
                in_block = False
                i += 2
                continue
            i += 1
            continue
        if in_str:
            code += l[i]
            if l[i] == "'":
                in_str = False
            i += 1
            continue
        if l[i] == "'":
            in_str = True
            code += l[i]
            i += 1
            continue
        if two == '/*':
            in_block = True
            i += 2
            continue
        if two == '--':
            cm = l[i:]
            break
        code += l[i]
        i += 1
    return code, cm, in_block


def tokenize(a, b):
    buf, anchors, in_block = [], {}, False
    for l in lines[a:b]:
        code, cm, in_block = strip_line(l, in_block)
        base = sum(len(x) for x in buf)
        m = re.search(r'--\s*P1\s+([\d.]+)', cm)
        if m and code.strip():
            anchors[base + len(code)] = 'P1 ' + m.group(1)
        buf.append(code + '\n')
    full = ''.join(buf)
    up = full.upper()
    toks, cur, pd, cd, i, st = [], '', 0, 0, 0, 0
    while i < len(full):
        if up[i:i + 4] == 'CASE' and (i == 0 or not (up[i - 1].isalnum() or up[i - 1] == '_')) and not up[i + 4:i + 5].isalnum():
            cd += 1
            cur += full[i:i + 4]
            i += 4
            continue
        if up[i:i + 3] == 'END' and (i == 0 or not (up[i - 1].isalnum() or up[i - 1] == '_')) and not (up[i + 3:i + 4].isalnum() or up[i + 3:i + 4] == '_'):
            cd = max(0, cd - 1)
            cur += full[i:i + 3]
            i += 3
            continue
        ch = full[i]
        if ch == '(':
            pd += 1
        elif ch == ')':
            pd -= 1
        if ch == '|' and full[i:i + 2] == '||' and pd == 0 and cd == 0:
            toks.append((st, i, cur))
            i += 2
            cur = ''
            st = i
            continue
        cur += ch
        i += 1
    if cur.strip():
        toks.append((st, len(full), cur))
    out = []
    for s, e, t in toks:
        t = re.sub(r'^\s*select\s+', '', t, flags=re.I)
        for part in re.split(r'\bas\s+lignedetail\d\s*,?', t, flags=re.I):
            p = part.strip().rstrip(',').strip()
            if not p or re.match(r'^(from|where)\b', p, re.I):
                continue
            out.append({'raw': p, 's': s, 'e': e})
    for pos, rf in sorted(anchors.items()):
        cand = [o for o in out if o['e'] <= pos]
        if cand:
            cand[-1]['ref'] = rf
    for o in out:
        o.setdefault('ref', None)
    return out


FIXED = {'F_FORMAT_TAUX': 10, 'F_FORMAT_MONTANT_BIS2': 19,
         'F_FORMAT_MONTANT_NEGATIF_19': 19, 'F_FORMAT_MONTANT': 19}


def split_args(s):
    d, out, cur = 0, [], ''
    for ch in s:
        if ch == '(':
            d += 1
        elif ch == ')':
            d -= 1
        if ch == ',' and d == 0:
            out.append(cur)
            cur = ''
            continue
        cur += ch
    out.append(cur)
    return out


def width(e):
    t = re.sub(r'\s+', ' ', e).strip().rstrip('|').strip()
    if not t:
        return None
    U = t.upper()
    if U.startswith('CASE'):
        ws_ = []
        for m in re.finditer(r'\b(THEN|ELSE)\b', U):
            rest = t[m.end():]
            RU = rest.upper()
            d = cd2 = j = 0
            out = ''
            while j < len(rest):
                if RU[j:j + 4] == 'CASE':
                    cd2 += 1
                if RU[j:j + 3] == 'END' and not RU[j + 3:j + 4].isalnum():
                    if cd2 == 0:
                        break
                    cd2 -= 1
                if d == 0 and cd2 == 0 and re.match(r'\b(WHEN|ELSE)\b', RU[j:]):
                    break
                if rest[j] == '(':
                    d += 1
                elif rest[j] == ')':
                    d -= 1
                out += rest[j]
                j += 1
            w = width(out)
            if w:
                ws_.append(w)
        return max(ws_) if ws_ else None
    # SUBSTR(expr, inicio, n) -> n  (tem de vir ANTES das funcoes de formato:
    # o SUBSTR corta o resultado de F_FORMAT_* e e ele que manda na largura)
    m = re.match(r'^SUBSTR\s*\((.*)\)$', t, re.I)
    if m:
        parts = split_args(m.group(1))
        if len(parts) >= 3:
            try:
                return int(parts[2].strip())
            except Exception:
                pass
    m = re.match(r'^[RL]PAD\s*\((.*)\)$', t, re.I)
    if m:
        parts = split_args(m.group(1))
        if len(parts) >= 2:
            try:
                return int(parts[1].strip())
            except Exception:
                pass
    for fn, w in FIXED.items():
        if re.search(r'\b' + fn + r'\s*\(', t, re.I):
            return w
    if re.fullmatch(r"'[^']*'", t):
        return len(t) - 2
    if ':MASYSDATE' in U:
        return 12
    if re.search(r"TO_CHAR\s*\(.*'YYYYMMDD'", t, re.I):
        return 8
    m = re.match(r"^NVL\s*\((.*)\)$", t, re.I)
    if m:
        parts = split_args(m.group(1))
        if len(parts) == 2:
            d = parts[1].strip()
            if re.fullmatch(r"'[^']*'", d):
                return len(d) - 2
    return None


toks = tokenize(589, 1068)
unk = [t['raw'][:50] for t in toks if width(t['raw']) is None]
tot = sum(width(t['raw']) or 0 for t in toks)
print('spool #1   : %d tokens, %d octetos calculados, %d sem largura'
      % (len(toks), tot, len(unk)))
print('delta spool vs regua V44 : %d' % (tot - TOT_V44))
if unk:
    print('tokens sem largura:')
    for u in unk[:10]:
        print('   ?', u)

# ------------------------------------------------------- alinhamento
pos = 0
ok = bad = 0
erros = []
inferidas = []
for t in toks:
    w = width(t['raw'])
    if w is None:
        # largura desconhecida : deduz-se do campo da notice nesta posicao
        nf = next((f for f in v44 if f['start'] == pos), None)
        w = nf['len'] if nf else 0
        inferidas.append((t['raw'][:45], w, nf['ref'] if nf else '?'))
    campos = [f for f in v44 if f['start'] < pos + w and f['start'] + f['len'] > pos]
    t['campos'] = [f['ref'] for f in campos]
    if t.get('ref'):
        # ancora de BLOCO (ex. 'P1 25') cobre os campos 'P1 25.x'
        pref = t['ref'] + '.'
        if t['ref'] in t['campos'] or any(c.startswith(pref) for c in t['campos']):
            ok += 1
        else:
            bad += 1
            erros.append((pos, t['ref'], t['campos'][:3], t['raw'][:45]))
    pos += w
print()
print('ANCORAS: %d certas / %d erradas' % (ok, bad))
if inferidas:
    print('larguras inferidas:')
    for r, w, rf in inferidas:
        print('   %-46s -> %d  (%s)' % (r, w, rf))
for e in erros[:12]:
    print('   X off=%-5d ancora=%-10s alinhou=%s | %s' % e)
