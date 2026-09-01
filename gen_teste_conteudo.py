"""Gera teste_conteudo.sql : compara a tabela com o ficheiro REAL, campo a campo.

O CRRCORP_P1 e o ficheiro que o spool produz hoje. Para alguns engajamentos
concretos, este script le o valor de cada campo no ficheiro (pela posicao da
regua V44), converte-o ao valor tipado e escreve-o no SQL como valor esperado.
O script resultante devolve, por engajamento, SO as colunas que divergem.

Nao precisa do spool vPACT: valida ja se os 261 mapeamentos estao certos.
"""
import io
import os
import re
import sys

import openpyxl

REAL = (r'C:/Users/elder/AppData/Local/Temp/claude/'
        r'C--Users-elder-Documents/c833d650-7531-40f4-951e-02752213f565/'
        r'scratchpad/CRRCORP_P1.txt')
N_ENG = 5

# Colunas que dependem da EXECUCAO, nao dos dados: divergem sempre porque o
# ficheiro foi gerado noutro momento. Comparar so criaria ruido.
EXCLUIR = {'P1_H_0_5'}   # horodatage de tratamento (p_masysdate)
NL = chr(10)
Q = chr(39)

# ---------------------------------------------------------------- notice
ws = openpyxl.load_workbook('Notice PACTV4.5_v1.0.xlsx', data_only=True)['PACT Corp']
FMT, DEC = {}, {}
for r in range(4, ws.max_row + 1):
    if str(ws.cell(r, 1).value or '').strip() != 'P1':
        continue
    ref = str(ws.cell(r, 5).value or '').strip()
    FMT[ref] = str(ws.cell(r, 20).value or '').strip().upper()
    regra = str(ws.cell(r, 21).value or '')
    m = re.search(r'(\d+)\s*d.cimale', regra, re.I)
    DEC[ref] = int(m.group(1)) if m else 0

# ------------------------------------------------- regua + tokens (variante 1)
src = open('align_v44.py', encoding='utf-8').read()
src = src.split('# ------------------------------------------------------- alinhamento')[0]
ns = {}
buf = io.StringIO()
_o = sys.stdout
sys.stdout = buf
exec(src, ns)
sys.stdout = _o
tokenize, width, v44 = ns['tokenize'], ns['width'], ns['v44']

DDL = open('ENG_CORP_P1_BIS.sql', encoding='utf-8').read()
TIPOS = {}
for m in re.finditer(r'^\s+(P1_[A-Z0-9_]+)\s+(NUMBER\((\d+)(?:,\s*(\d+))?\)|DATE|VARCHAR2\((\d+)\))',
                     DDL, re.M):
    TIPOS[m.group(1)] = m.group(2)

# colunas realmente alimentadas pela procedure (INSERT #1)
proc = open('pack_alim_tab_envoi_crrv4_P_ALIM_ENG_CORP_P1_BIS.sql', encoding='utf-8').read()
bloco1 = proc.split('-- INSERT #1')[1].split('-- INSERT #2')[0]
ALIMENTADAS = set(re.findall(r'AS (P1_[A-Z0-9_]+)', bloco1))


def col_de(ref):
    if '(P1)' in ref:
        return 'P1_H_' + ref.replace('(P1)', '').strip().replace('.', '_')
    return 'P1_' + ref.split()[1].replace('.', '_')


# posicao de cada coluna, percorrendo a variante 1
campos = []
pos = 0
for t in tokenize(589, 1068):
    w = width(t['raw'])
    if w is None:
        nf = next((f for f in v44 if f['start'] == pos), None)
        w = nf['len'] if nf else 0
    achados = [f for f in v44 if f['start'] < pos + w and f['start'] + f['len'] > pos]
    if len(achados) == 1:
        c = col_de(achados[0]['ref'])
        if c in ALIMENTADAS and c in TIPOS:
            campos.append((c, pos, w, achados[0]['ref']))
    pos += w
# dedup, fica a 1a posicao de cada coluna
vistos, CAMPOS = set(), []
for c, o, w, ref in campos:
    if c not in vistos:
        vistos.add(c)
        CAMPOS.append((c, o, w, ref))
print('colunas comparaveis (INSERT #1): %d' % len(CAMPOS))

# ---------------------------------------------------------------- ficheiro
if not os.path.exists(REAL):
    raise SystemExit('ficheiro real nao encontrado: ' + REAL)
ID_OFF, ID_W = 180, 40
linhas = []
with open(REAL, 'r', encoding='latin-1') as f:
    for i, l in enumerate(f):
        l = l.rstrip('\r\n')
        if l[ID_OFF:ID_OFF + ID_W].strip():
            linhas.append(l)
        if len(linhas) >= N_ENG:
            break


def esperado(valor, ref, tipo):
    """Valor do ficheiro -> literal SQL do valor tipado esperado."""
    v = valor.strip()
    if not v:
        return None
    if tipo == 'DATE':
        return "TO_DATE(%s%s%s,%sYYYYMMDD%s)" % (Q, v, Q, Q, Q) if re.fullmatch(r'\d{8}', v) else False
    if tipo.startswith('NUMBER'):
        m = re.fullmatch(r'([+-]?)(\d+)', v)
        if not m:
            return False
        sinal, dig = m.group(1), m.group(2)
        d = DEC.get(ref, 0)
        num = ('-' if sinal == '-' else '') + (dig[:-d] or '0') + ('.' + dig[-d:] if d else '')
        return str(float(num))
    return Q + v.replace(Q, Q + Q) + Q


partes = []
for l in linhas:
    idv = l[ID_OFF:ID_OFF + ID_W].strip()
    casos = []
    for c, off, w, ref in CAMPOS:
        if c in EXCLUIR:
            continue
        tipo = TIPOS[c]
        exp = esperado(l[off:off + w], ref, tipo)
        if exp is False:
            continue
        if exp is None:
            casos.append("         CASE WHEN %s IS NOT NULL THEN %s%s(esperado NULL) %s END ||"
                         % (c, Q, c, Q))
        elif tipo == 'DATE':
            casos.append("         CASE WHEN %s IS NULL OR TRUNC(%s) <> %s THEN %s%s %s END ||"
                         % (c, c, exp, Q, c, Q))
        elif tipo.startswith('NUMBER'):
            d = DEC.get(ref, 0)
            casos.append("         CASE WHEN %s IS NULL OR ROUND(%s,%d) <> %s THEN %s%s %s END ||"
                         % (c, c, d, exp, Q, c, Q))
        else:
            casos.append("         CASE WHEN %s IS NULL OR RTRIM(%s) <> %s THEN %s%s %s END ||"
                         % (c, c, exp, Q, c, Q))
    if not casos:
        continue
    corpo = NL.join(casos).rstrip('|').rstrip()
    partes.append("  SELECT %s%s%s AS engajamento," % (Q, idv, Q) + NL + corpo + NL
                  + "           AS colunas_divergentes" + NL
                  + "    FROM ENG_CORP_P1_BIS" + NL
                  + "   WHERE P1_H_1_11 = %s%s%s" % (Q, idv, Q))

sql = [
    "-- Compara a ENG_CORP_P1_BIS com o ficheiro CRRCORP_P1 realmente gerado.",
    "-- Os valores esperados foram lidos do ficheiro pela posicao da regua V44",
    "-- e convertidos ao tipo da coluna. Devolve SO as colunas que divergem:",
    "-- coluna vazia = engajamento totalmente conforme.",
    "--",
    "-- ATENCAO: o ficheiro e uma FOTOGRAFIA (arrete 20250630). Se os dados de",
    "-- DEV2 mudaram desde entao, parte das divergencias e legitima. O sinal a",
    "-- procurar e o PADRAO: uma coluna que diverge em TODOS os engajamentos e",
    "-- suspeita de mapeamento errado; divergencias dispersas sao dados que mudaram.",
    "--",
    "-- Divergencias esperadas (ja conhecidas, ver docs/SIRL-1224.md):",
    "--   P1_3_20 e P1_31_17/18 : o spool TRUNCA o valor no ficheiro,",
    "--   a tabela guarda o valor correto. Divergir aqui e o comportamento certo.",
    "SET LINESIZE 32000",
    "COLUMN engajamento        FORMAT A26",
    "COLUMN colunas_divergentes FORMAT A120",
    "",
    (NL + "UNION ALL" + NL).join(partes) + ";",
]
open('teste_conteudo.sql', 'w', encoding='utf-8').write(NL.join(sql) + NL)
print('engajamentos comparados : %d' % len(partes))
print('-> teste_conteudo.sql')
